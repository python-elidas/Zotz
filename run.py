import openpyxl as xls
import simply_sqlite as SQL
from datetime import datetime

now = datetime.now()
today = now.strftime('%x')
file = 'C:/Users/osgum/Desktop/Zotz/2021-Gastos MAKRO.xlsx'
db = SQL.SQL('files/zotz_db')

wb = xls.load_workbook(filename=file, read_only=False)
sheets = wb.sheetnames[1:10]

for sheet in sheets:
    for row in wb[sheet].iter_rows(min_row=61):
        info = db.show_all_rows('Articulos')
        if not row[1].value is None:
            db.insert_info('Articulos', 'Codigo', row[1].value)
            db.update('Articulos', 'Descripcion', row[1].value, row[2].value)
            db.update('Articulos', 'Tipo', row[1].value, row[3].value)
            db.update('Articulos', 'TimeStamp', row[1].value, today)
    print(info)
