import openpyxl as xls
import simply_sqlite as SQL
from datetime import datetime

now = datetime.now()
today = now.strftime('%x')
file = 'C:/Users/osgum/Desktop/Zotz/2021-Gastos MAKRO.xlsx'
db = SQL.SQL('files/zotz_db')

wb = xls.load_workbook(filename=file, read_only=False)
sheets = wb.sheetnames[1:10]
codes = list()

for sheet in sheets:
    for row in wb[sheet].iter_rows(min_row=61):
        info = db.show_all_rows('Articulos')
        if not row[1].value is None and\
                not row[1].value in codes:
            db.insert_info('Articulos', 'Codigo', row[1].value)
            db.update(
                'Articulos', 'Descripcion', 'Codigo',
                (row[2].value, row[1].value))
            db.update(
                'Articulos', 'Tipo', 'Codigo',
                (row[3].value, row[1].value))
            db.update(
                'Articulos', 'TimeStamp', 'Codigo',
                (today, row[1].value))
            codes.append(row[1].value)
    print(info)
