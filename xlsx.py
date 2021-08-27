'''
Author: Elidas
Email: pyro.elidas@gmail.com
Python version: 3.9.1
Date: 2021-08-26T11:18:58.589Z
Version: 0.0.1
'''

# __LYBRARIES__ #
import openpyxl as xls
import simply_sqlite as SQL
from makro import makro
from datetime import datetime
from main import Main_Window

# __MAIN CODE__ #


def m_xlsx(xcel, pdf):
    today = datetime.now().strftime('%x')
    # accedemos a la base de datos
    db = SQL.SQL('files/zotz_db')
    # accedemos al excel
    wb = xls.load_workbook(filename=xcel, read_only=False)
    # obtenemos el nombre de la factura y su información
    new, bill = makro(pdf)
    # creamos la hoja con la que trabajaremos
    ws = wb.copy_worksheet(wb['Siguiente'])
    # Establecemos el tituo de la nueva hoja
    ws.title = new
    # Comenzamos a introducir infromacion en la hoja
    ws['G2'] = bill['num'][0]
    ws['G3'] = bill['num'][1]
    ws['G4'] = bill['fecha']

    items = bill['articulos']
    row = 62
    for item in items:
        ws[f'A{row}'] = row - 61
        ws[f'B{row}'] = item['codigo']
        ws[f'C{row}'] = item['desc']
        if len(db.show_one_row('Articulos', 'Codigo', item['codigo'])) == 0:
            # Main_Window.new_item(item)
            ws[f'D{row}'] = 'PNDT'
        else:
            ws[f'D{row}'] = db.show_one_row(
                'Articulos', 'Codigo', item['codigo'])[0][2]
        ws[f'E{row}'] = item['prec ud']
        ws[f'F{row}'] = item['ud pac']
        ws[f'G{row}'] = item['precio']
        ws[f'H{row}'] = item['uds']
        ws[f'I{row}'] = item['precio'] * int(item['uds'])
        ws[f'J{row}'] = item['iva']

        # insertamos la siguiente fila
        ws.insert_rows(row+1)
        # cpiamos las formulas pertinentes
        ws[f'K{row+1}'] = str(ws[f'K{row}'].value).replace(str(row), str(row+1))
        ws[f'L{row+1}'] = ws[f'L{row}'].value.replace(str(row), str(row+1))
        ws[f'M{row+1}'] = ws[f'M{row}'].value.replace(str(row), str(row+1))
        # copaimos los formatos de las celdas
        x = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
        for j in range(13):
            ws[f'{x[j]}{row+1}']._style = ws[f'{x[j]}{row}']._style
        # avanzamos a la siguiente fila
        row += 1

    if len(bill['descuentos']) != 0:
        for item in bill['descuentos']:
            # insertamos la información
            ws[f'A{row}'] = row - 61
            ws[f'B{row}'] = item['code']
            ws[f'C{row}'] = 'Descuento'
            ws[f'D{row}'] = 'MP'
            ws[f'E{row}'] = item['val']
            ws[f'F{row}'] = 1
            ws[f'G{row}'] = item['val']
            ws[f'H{row}'] = 1
            ws[f'I{row}'] = item['val']
            ws[f'J{row}'] = item['iva']

            # insertamos la siguiente fila
            ws.insert_rows(row+1)
            # cpiamos las formulas pertinentes
            ws[f'K{row+1}'] = str(ws[f'K{row}'].value).replace(str(row), str(row+1))
            ws[f'L{row+1}'] = ws[f'L{row}'].value.replace(str(row), str(row+1))
            ws[f'M{row+1}'] = ws[f'M{row}'].value.replace(str(row), str(row+1))
            # copaimos los formatos de las celdas
            x = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
            for j in range(13):
                ws[f'{x[j]}{row+1}']._style = ws[f'{x[j]}{row}']._style
            # avanzamos a la siguiente fila
            row += 1

    # eliminamos la fila que sobra
    ws.delete_rows(row)

    # ponemos en orden el desorden
    ws[f'F{row+2}'].value = str(ws[f'F{row+2}'].value)\
        .replace('F62', f'F{row-1}')
    ws[f'H{row+2}'].value = str(ws[f'H{row+2}'].value)\
        .replace('H62', f'H{row-1}')
    ws[f'I{row+2}'].value = str(ws[f'I{row+2}'].value)\
        .replace('I62', f'I{row-1}')
    ws[f'L{row+2}'].value = str(ws[f'L{row+2}'].value)\
        .replace('L62', f'L{row-1}')
    ws[f'M{row+2}'].value = str(ws[f'M{row+2}'].value)\
        .replace('M62', f'M{row-1}')

    # Guardamos los cambios
    wb.save(xcel)


if __name__ == '__main__':
    xcel = 'C:/Users/osgum/Desktop/Zotz/2021-Gastos MAKRO.xlsx'
    pdf = '///Users/osgum/Desktop/Zotz/Facturas_MAKRO/21-05-08-MAKRO-01.pdf'
    m_xlsx(xcel, pdf)

# __NOTES__ #
'''

'''

# __BIBLIOGRAPHY__ #
'''
    pypi:       https://pypi.org/project/openpyxl/
    openpyxl:   https://openpyxl.readthedocs.io/en/stable/
'''
