'''
Author: Elidas
Email: pyro.elidas@gmail.com
Python version: 3.9
Date: 2021-08-23
Version: 1.3.5
'''

# __LIBRARIES__ #
from tkinter import *
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar
import openpyxl as xls
from siguiente import Excel
import threading as th
from file_manage import *


# __MAIN CODE__ #
class Main_Window(Tk):
    def __init__(self):
        Tk.__init__(self)
        
        self.ind = 0
        self.n, self.m = 1, 2
        
        self.exe = th.Thread(target=self.run)

        self._frame = Main_Frame(self)
        self._frame.pack(fill=NONE, expand=1)

        self.title('Contabilidad Zotz')
        self.iconbitmap('files/logo.ico')
        self.geometry('600x250')
        self.resizable(FALSE, FALSE)
        
        self.exit = Button(self, text = 'Salir', command = self.out)
        self.exit.pack(anchor = SE, side = BOTTOM, padx = 10, pady = 5)
    
    def out(self):
        self.kill_thread()
        self.destroy()
    
    def kill_thread(self):
        if self.exe.is_alive():
            self.exe.join()
        
    def run(self):
        # modificamos la ventana para adaptarse a la nueva información
        self.geometry('600x275')
        
        # Obtenemos la información del campo y deshabilitamos el campo y botón
        dir = self._frame.folder.get()
        self._frame.folder['state'] = 'disabled'
        self._frame.f_b['state'] = 'disabled'
        
        # Obtenemos la información del campo y deshabilitamos el campo y botón
        bills = file.listdir(dir)
        excel = f'{self._frame.excel.get()}/{get_xcl_name(dir)}.xlsx'
        print(excel)
        self._frame.excel['state'] = 'disabled'
        self._frame.f_e['state'] = 'disabled'
        
        # Empezamos a mostrar nueva información y creamos la barra de progreso
        Label(self._frame, text='Procesando Archivos...')\
            .grid(row=2, column=0, sticky=W)
        self.p_b = Progressbar(self._frame, length=500, )
        self.p_b.grid(row=3, columnspan=4)
        
        # Comprobamos que facturas han sido pasadas para no repetir
        process = list()
        for bil in bills:
            nme = bil.split('.')[0].split(' - ')[-1]
            if '.pdf' in nme:
                nme = nme.replace('.pdf', '')
            if '.pdf' in bil and not nme in self._frame.ws_nms:
                process.append(bil)
                
        # Establecemos el maximo de la barra de progreso
        self.p_b.config(maximum=len(process))
        
        # Empezamos a leer y convertir información
        try:
            self.n, self.m = 1, len(process)
            # Iteramos por todas las facturas
            for self.bil in process:
                # adecuamos el nombre del archivo
                pdf = dir.replace('C:', '//') + '/' + self.bil
                # mostramos mas infromacion
                Label(self._frame,
                      text=f'{self.n} de {self.m}')\
                    .grid(row=2, column=2)                    
                Label(self._frame, 
                        text=f'Archivo {self.bil}')\
                    .grid(row=4, column=0, columnspan=3, sticky=W)
                Label(self._frame,
                        text='\tLeyendo Archivo PDF')\
                    .grid(row=6, column=0, sticky=W)
                Label(self._frame,
                        text='\tCreando Hoja Excel')\
                    .grid(row=7, column=0, sticky=W)
                Label(self._frame,
                        text='\tEscribiendo Información')\
                    .grid(row=8, column=0, sticky=W)
                Excel(excel, pdf, self)
                self.n += 1
                self.p_b.step(1)
                # time.sleep(2)
            # Una vez finalizado, avisamos.
            messagebox.showinfo(
                message="El archivo Excel ha sido actualizado.\n reinicie el programa para procesar mas archivos.",
                title="Porceso completado.")
            self._frame.exe.config(text='Apagar', command=self.out)
        # si esta el excel abierto
        except PermissionError:
            messagebox.showerror(
                message="Cierra el fichero Excel y vuelve a intentarlo",
                title="Permission Error"
            )
            self.run()
        # si se produce cualquier error
        except Exception as e:
            f, line, func = self.catch_error(e)
            messagebox.showerror(
                title=type(e).__name__,
                message=f'Ha ocurrido un error:\
                \n[+] En el archivo {self.bil}\
                \n[+] En el script {f}\
                \n[+] En la linea {line.split(" ")[-1]}\
                \n[+] En la funcion {func}\
                \nSe almacenará en el Registro de errores')
            self.record_errors(e)
        
    def catch_error(self, e):
        import traceback
        
        info = str(traceback.extract_tb(e.__traceback__)[-1])\
            [1:-1].replace('\\', '/').split(',')
        return  info[0].split('/')[-1],\
                info[-1].split(' in ')[0],\
                info[-1].split(' in ')[-1]
            
    def record_errors(self, e):
        import traceback, datetime
        
        error = str(traceback.format_exc()).split('\n')[1:-1]
        now = datetime.datetime.now()
        day = now.strftime('%d/%m/%Y %H:%M')
        
        with open('files/ErrorLog.txt', 'a') as fw:
            fw.write(f'Error at date {day}:\n')
            fw.write(f'[+] While reading the file: {self.bil}\n')
            fw.write(f'[+] The type of the error was: {type(e).__name__}\n')
            fw.write(f'The text was:\n')
            for info in error:
                if info.startswith('  '):
                    fw.write(f'[+] {str(info)[2:]}\n')
                else:
                    fw.write(f'[+]\t{str(info)}\n')
            fw.write('\n')
            fw.close()
        
    def switch_frames(self, frame):
        if self._frame is not None:
            self._frame.destroy()
        self._frame = frame(self)
        self._frame.pack(fill=NONE, expand=1)
        

class Main_Frame(Frame):
    def __init__(self, master):
        Frame.__init__(self, master)

        self._master = master

        Label(self, text='Selecciona la carpeta de las facturas')\
            .grid(row=0, column=0, sticky=W)
        self.folder = Entry(self, width=30)
        self.folder.grid(row=1, column=0, sticky=W)
        self.f_b = Button(self, text='···', width=3, command=self.select_folder)
        self.f_b.grid(row=1, column=1)

        Label(self, text='Selecciona la carpeta del Excel')\
            .grid(row=0, column=2, sticky=W, padx=10)
        self.excel = Entry(self, width=35)
        self.excel.grid(row=1, column=2, sticky=W, padx=10)
        self.f_e = Button(self, text='···', width=3, command=self.select_excel)
        self.f_e.grid(row=1, column=3, sticky=E)

        self.exe = Button(self, text='Run', command=self.exe_run)
        self.exe.grid(row=2, column=3, sticky=E, pady=10)

    def select_excel(self):
        xcel = filedialog.askdirectory(initialdir=self._init)
        self.excel.delete(0, END)
        self.excel.insert(0, xcel)
        if not find_excel(self.excel.get(), self.folder.get()):
            nme = create_new_excl(self.excel.get(), self.file.get())
        else:
            nme = get_xcl_name(self.folder.get())
        wb = xls.load_workbook(f'{self.excel.get()}/{nme}.xlsx', read_only=True)
        self.ws_nms = list(wb.sheetnames)

    def select_folder(self):
        import os as file
        dir = filedialog.askdirectory()
        self.folder.delete(0, END)
        self.folder.insert(0, dir)
        same = messagebox.askyesno(
            title='Selección de carpeta',
            message='¿Desea emplear la misma carpeta para guardar el fichero Excel?'
        )
        if same:
            self._init = dir
            self.excel.delete(0, END)
            self.excel.insert(0, dir)
            self.excel['state'] = 'disabled'
            self.f_e['state'] = 'disabled'
            if not find_excel(self.excel.get(), self.folder.get()):
                nme = create_new_excl(self.excel.get(), self.folder.get())
            else:
                nme = get_xcl_name(self.folder.get())
            wb = xls.load_workbook(f'{self.excel.get()}/{nme}.xlsx', read_only=True)
            self.ws_nms = list(wb.sheetnames)
        else:
            self._init = file.getcwd()
            
    def exe_run(self):
        if not self.excel.get() == '' and not self.folder.get() == '':
            self._master.exe.start()
        else:
            messagebox.showerror(
                message="Verifique que ambos campos están complimentados y vuelva a intentarlo",
                title='Campo vacío'
            )



def run():
    main = Main_Window()
    main.mainloop()


if __name__ == '__main__':
    run()

# __NOTES__ #
'''
Creación del exe:
    pyinstaller --noconfirm --onedir --windowed --icon "C:/Users/osgum/github/Zotz/files/logo.ico" --name "Zotz_Cont v1.3.0" --add-data "C:/Users/osgum/github/Zotz/files;files/" --add-data "C:/Users/osgum/github/Zotz/LICENSE;." --add-data "C:/Users/osgum/github/Zotz/makro.py;." --add-data "C:/Users/osgum/github/Zotz/mercadona.py;." --add-data "C:/Users/osgum/github/Zotz/README.md;." --add-data "C:/Users/osgum/github/Zotz/sel_type.py;." --add-data "C:/Users/osgum/github/Zotz/siguiente.py;."  "C:/Users/osgum/github/Zotz/main.py"
'''

# __BIBLIOGRAPHY__ #
'''
    · ProgressBar: https://recursospython.com/guias-y-manuales/barra-de-progreso-progressbar-tcltk-tkinter/

'''
