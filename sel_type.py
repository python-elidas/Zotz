'''
Author: Elidas
Email: pyro.elidas@gmail.com
Python version: 3.9.1
Date: 7/12/2021
Version: 1.0.0
'''

# __LIBRARIES__ #
from os import kill
from tkinter import *
from tkinter.ttk import Combobox
import openpyxl as xls
import simply_sqlite as SQL
from datetime import datetime


# __MAIN CODE__ #
class Sel_Type(Tk):
    def __init__(self, no_ID, wb):
        Tk.__init__(self)
        
        # variables
        self.no_ID = no_ID
        self.ws = wb['Datos']
        self.ind = 0
        self.item = self.no_ID[0]
        
        # Elementos generales de la ventana
        self.title('Elemento NO hallado')
        self.iconbitmap('files/logo.ico')
        self.geometry('350x200')
        self.resizable(FALSE, FALSE)
        
        # Entradas y editables
        Label(self, text='Por Favor, selecciona el tipo de gasto:')\
            .pack()
        self.name = Label(self, text=self.no_ID[0]['desc'])
        self.name.pack()
        self.id = Combobox(
            self,
            values=self.get_id())
        self.id.pack()
        if not self.ind == len(self.no_ID)-1:
            self.save = Button(self, text='Siguiente', command=self.next)
        else:
            self.save = Button(self, text='Guardar', command=self.save_ID)
        self.save.pack()  
        
    def next(self):
        # es mas sencillo esto que cambiar toda la funcion
        self.save_ID()
        self.ind += 1
        self.item = self.no_ID[self.ind]
        self.name.config(text=self.item['desc'])
        if self.ind == len(self.no_ID)-1:
            self.save.config(text='Guardar', command=self.save_ID)
        
    def get_id(self):
        self.IDs = dict()
        item, row = True, 27
        while not item == None:
            key = self.ws[f'E{row}'].value
            value = self.ws[f'D{row}'].value
            self.IDs[key] = value
            row += 1
            item = self.ws[f'E{row}'].value
        return list(self.IDs.keys())
    
    def save_ID(self):
        now = datetime.now()
        today = now.strftime('%x')
        db = SQL.SQL('files/zotz_db')
        #Insertamos el código del articulo
        db.insert_info('Articulos', 'Codigo', self.item['codigo'])
        # insertamos la descripcion
        db.update(
            'Articulos', 'Descripcion', 'Codigo',
            (self.item['desc'], self.item['codigo']))
        # Insertamos el ID
        db.update(
            'Articulos', 'Tipo', 'Codigo',
            (self.IDs[self.id.get()], self.item['codigo']))
        # Insertamos la fecha
        db.update(
            'Articulos', 'TimeStamp', 'Codigo',
            (today, self.item['codigo']))
        if self.ind == len(self.no_ID)-1:
            self.destroy()

# __RUN CODE__ #
if __name__ == '__main__':
    xcel = 'C:/Users/osgum/Desktop/Zotz/Facturas_MAKRO/Test_1/2021-Gastos MAKRO respaldo.xlsx'
    wb = xls.load_workbook(filename=xcel, read_only=False)
    main = Sel_Type(wb=wb)
    main.mainloop()


#! __NOTES__ !#
'''
    
'''

#? __BIBLIOGRAPHY__ ?#
'''

'''