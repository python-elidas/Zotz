'''
Author: Elidas
Email: pyro.elidas@gmail.com
Python version: 3.9
Date: 2021-08-26
Version: 1.3.5
'''

# __LYBRARIES__ #
from tkinter import *
import openpyxl as xls
import simply_sqlite as SQL
from makro import Makro
from mercadona import Mercadona
from datetime import datetime
from sel_type import Sel_Type
import time
from auxiliares.toolPrint import *

# __MAIN CODE__ #
class Excel:
    def __init__(self, xcel, pdf, master = ''):
        # obtiene la fecha de hoy
        self.today = datetime.now().strftime('%x')
        # obtenemos el nombre de la factura y su información
        self.new, self.bill = self.select_case(pdf)
        # damos Información al usuario:
        try:
            a = Label(master._frame, text='OK')
            a.grid(row=5, column=1)
        except:
            pass
        # accedemos al excel
        self.wb = xls.load_workbook(filename=xcel, read_only=False)
        # Comprobemos que el año es correcto:
        if str(self.wb['Datos']['L16'].value).upper() == 'YYYY':
            self.wb['Datos']['L16'].value = \
                f"20{pdf.split('/')[-1].split(' - ')[0].split('-')[0]}"
        # creamos la hoja con la que trabajaremos
        self.ws = self.wb.copy_worksheet(self.wb['Siguiente'])
        # Establecemos el tituo de la nueva hoja
        self.ws.title = self.new
        # damos información al usuario:
        try:
            b = Label(master._frame, text='OK')
            b.grid(row=6, column=1)
        except:
            pass
        # Escribimos la cabecera
        self.write_head()
        # Comprobamos si existen elementos sin ID
        self.get_id()
        # Escribimos los articulos:
        self.write_items()
        # Escribimos los descuentos si existen:
        try:
            self.write_discounts()
        except KeyError:
            pass
        # reordenamos el desorden:
        self.reorder() 
        # ejecutamos el siguiente paso:
        self.overview()
        # Guardamos los cambios
        self.wb.save(xcel)
        # damos información al usuario:
        try:
            c = Label(master._frame, text='OK')
            c.grid(row=7, column=1)
        except:
            pass
        print(f'Bill {self.new} Saved correctly!')
        time.sleep(1.5)
        try:
            a.destroy()
            b.destroy()
            c.destroy()
        except: 
            pass
        
    def select_case(self, pdf):
        case = {
            'Makro' : Makro(pdf),
            'Mercadona' : Mercadona(pdf)
        }
        self.prov = pdf.split(' - ')[1]
        return case[self.prov].result()

    def write_head(self): # Escribimos los datos relevantes de la factura
        # número de factura y fecha
        try:
            self.ws['G2'] = self.bill['Factura'][0]
            self.ws['G3'] = self.bill['Factura'][1]
            self.ws['G4'] = self.bill['fecha']
        except KeyError:
            self.ws['D2'] = 'Fact. Dev.'
            self.ws['G2'] = self.bill['Factura devolucion'][0]
            self.ws['G3'] = self.bill['Factura devolucion'][1]
            self.ws['G4'] = self.bill['fecha']

    def get_id(self):
        db = SQL.SQL('files/zotz_db')
        no_ID = set()
        for item in self.bill['articulos']:
            db_item = db.show_one_row(  # Por comodidad
                self.prov,
                'Codigo',
                item['codigo'])
            if len(db_item) == 0 \
                and not item['codigo'] in no_ID:
                no_ID.add(item)
        # listPrint(no_ID)
        if not len(no_ID) == 0:
            id = Sel_Type(no_ID, self.wb, self.prov)
            id.mainloop()
            
    def write_items(self): # empezamos con los articulos:
        db = SQL.SQL('files/zotz_db')
        items = self.bill['articulos']
        self.row = 102
        for item in items:
            # obtenemos la referencia del tipo de producto
            self.ws[f'A{self.row}'] = self.row - 61  # escribimos el numero de la fila
            self.ws[f'B{self.row}'] = item['codigo']  # escribimos la referencia
            self.ws[f'C{self.row}'] = item['desc']  # escribimos la descripción
            self.ws[f'D{self.row}'] = list(db.show_one_row(
                self.prov, 'Codigo', item['codigo']))[0][2]
            self.ws[f'E{self.row}'] = item['prec ud']
            self.ws[f'F{self.row}'] = item['ud pac']
            self.ws[f'G{self.row}'] = item['precio']
            self.ws[f'H{self.row}'] = item['uds']
            self.ws[f'I{self.row}'] = item['precio'] * int(item['uds'])
            self.ws[f'J{self.row}'] = item['iva']
            
            #Insertamos la siguiente fila:
            self.insert_new_row()

    def insert_new_row(self):
            # insertamos la siguiente fila
            self.ws.insert_rows(self.row+1)
            # copiamos las formulas pertinentes
            self.ws[f'K{self.row+1}'] = str(self.ws[f'K{self.row}'].value)\
                .replace(str(self.row), str(self.row+1))
            self.ws[f'L{self.row+1}'] = self.ws[f'L{self.row}'].value\
                .replace(str(self.row), str(self.row+1))
            self.ws[f'M{self.row+1}'] = self.ws[f'M{self.row}'].value\
                .replace(str(self.row), str(self.row+1))
            # copaimos los formatos de las celdas
            cols = [
                'A', 'B', 'C',
                'D', 'E', 'F',
                'G', 'H', 'I',
                'J', 'K', 'L', 'M'
            ]
            for col in cols:
                self.ws[f'{col}{self.row+1}']._style = self.ws[f'{col}{self.row}']._style
            # avanzamos a la siguiente fila
            self.row += 1

    def write_discounts(self): # Pasamos a los descuentos
        if not len(self.bill['descuentos']) == 0:
            for item in self.bill['descuentos']:
                # insertamos la información
                self.ws[f'A{self.row}'] = self.row - 61
                self.ws[f'B{self.row}'] = item['code']
                self.ws[f'C{self.row}'] = 'Descuento'
                self.ws[f'D{self.row}'] = 'MP'  # No siempre
                self.ws[f'E{self.row}'] = item['val']
                self.ws[f'F{self.row}'] = 1
                self.ws[f'G{self.row}'] = item['val']
                self.ws[f'H{self.row}'] = 1
                self.ws[f'I{self.row}'] = item['val']
                self.ws[f'J{self.row}'] = item['iva']

                #Insertamos la siguiente fila:
                self.insert_new_row()

    def reorder(self):
        # eliminamos la fila que sobra
        self.ws.delete_rows(self.row)
        # ponemos en orden el desorden
        # print(self.row)
        self.ws[f'F{self.row+2}'].value = str(self.ws[f'F{self.row+2}'].value)\
            .replace('F102', f'F{self.row-1}')
        self.ws[f'H{self.row+2}'].value = str(self.ws[f'H{self.row+2}'].value)\
            .replace('H102', f'H{self.row-1}')
        self.ws[f'I{self.row+2}'].value = str(self.ws[f'I{self.row+2}'].value)\
            .replace('I102', f'I{self.row-1}')
        self.ws[f'L{self.row+2}'].value = str(self.ws[f'L{self.row+2}'].value)\
            .replace('L102', f'L{self.row-1}')
        self.ws[f'M{self.row+2}'].value = str(self.ws[f'M{self.row+2}'].value)\
            .replace('M102', f'M{self.row-1}')
        self.row -= 1
        # adecuamos las fórmulas pertinentes la infromacion que tenemos:
        R, r = 19, 8
        J = [':$I$102', ':$K$102', ':$D$102']
        L = [':$L$102', ':$K$102', ':$D$102']
        M = [':$M$102', ':$K$102', ':$D$102']
        while r <= 11: 
            for elem in J:
                self.ws[f'J{r}'].value = str(self.ws[f'J{r}'].value)\
                    .replace(elem, elem[:-3]+str(self.row))
            for elem in L:
                self.ws[f'L{r}'].value = str(self.ws[f'L{r}'].value)\
                    .replace(elem, elem[:-3]+str(self.row))
            for elem in M:
                self.ws[f'M{r}'].value = str(self.ws[f'M{r}'].value)\
                    .replace(elem, elem[:-3]+str(self.row))
            r += 1
                    
        while not self.ws[f'K{R}'].value == None:
            for elem in J:
                self.ws[f'J{R}'].value = str(self.ws[f'J{R}'].value)\
                    .replace(elem, elem[:-2]+str(self.row))
            for elem in L:
                self.ws[f'L{R}'].value = str(self.ws[f'L{R}'].value)\
                    .replace(elem, elem[:-2]+str(self.row))
            for elem in M:
                self.ws[f'M{R}'].value = str(self.ws[f'M{R}'].value)\
                    .replace(elem, elem[:-2]+str(self.row))
            R += 1

    def overview(self):
        self.ws = self.wb['Resumen']
        self.row = 60
        bil = '\'' + self.new + '\''
        # buscamos el grupo de celdas con el que trabajar:
        while not str(self.ws[f'B{self.row}'].value).startswith('=Sig'):
            self.row += 1
        # iteramos las columnas y modificamos lo que hay que modificar:
        for i in range(2):
            col = 1
            while self.ws.cell(row=self.row, column=col).value is not None:
                # modifiquemos lo pertinente:
                val = str(self.ws.cell(row=self.row+i, column=col).value)
                # print(val)
                if val.startswith('=Sig') or val.startswith('=IF(Sig'):
                    # print(f'Antes: {str(self.ws.cell(row=self.row+i, column=col).value)}')
                    self.ws.cell(row=self.row+i, column=col).value = str(
                        self.ws.cell(row=self.row+i, column=col).value)\
                        .replace('Siguiente', bil)
                    # print(f'Despues: {str(self.ws.cell(row=self.row+i, column=col).value)}')
                col += 1


if __name__ == '__main__':
    xcel = 'C:/Users/osgum/Desktop/Zotz/Facturas_MERCADONA/Test/2021-Gastos MAKRO respaldo.xlsx'
    pdf = '///Users/osgum/Desktop/Zotz/Facturas_MERCADONA/Test/21-01-09 - Mercadona - A-V2021-67132.pdf'
    test = Excel(xcel, pdf)
    print('Done!')

# __NOTES__ #
'''


'''

# __BIBLIOGRAPHY__ #
'''
    pypi:       https: // pypi.org/project/openpyxl/
    openpyxl:   https: // openpyxl.readthedocs.io/en/stable/
'''
